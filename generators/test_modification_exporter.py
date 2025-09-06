"""
Test Modification Exporter

Exports generated test step modifications to Excel format with the same structure as QTEST_STTM.xlsx.
Includes Action column to specify ADD/MODIFY/DELETE operations for QA team.
"""

import pandas as pd
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

from templates.step_templates import GeneratedTestStep


class TestModificationExporter:
    """Exports test step modifications to Excel format compatible with QTEST"""
    
    def __init__(self, output_dir: str = "reports", logger: Optional[logging.Logger] = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)
    
    def export_to_excel(self, generated_steps: List[GeneratedTestStep], 
                       base_test_case_data: Dict[str, Any],
                       filename_suffix: str = None) -> str:
        """Export generated test steps to Excel file with QTEST structure"""
        
        if not generated_steps:
            self.logger.warning("No generated steps to export")
            return ""
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = f"_{filename_suffix}" if filename_suffix else ""
        filename = f"test_modifications{suffix}_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        
        self.logger.info(f"Exporting {len(generated_steps)} test modifications to {output_path}")
        
        # Create DataFrame with QTEST structure
        df = self._create_qtest_dataframe(generated_steps, base_test_case_data)
        
        # Export to Excel with proper formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Test Modifications', index=False)
            
            # Apply formatting
            self._apply_excel_formatting(writer.book['Test Modifications'])
        
        self.logger.info(f"Successfully exported test modifications to {output_path}")
        return str(output_path)
    
    def _create_qtest_dataframe(self, generated_steps: List[GeneratedTestStep], 
                               base_test_case_data: Dict[str, Any]) -> pd.DataFrame:
        """Create DataFrame with QTEST column structure"""
        
        # QTEST columns: Name, Id, Description, Precondition, Test Step #, Test Step Description, Test Step Expected Result
        # Plus our Action column
        
        rows = []
        
        for step in generated_steps:
            row = {
                'Name': base_test_case_data.get('name', ''),
                'Id': base_test_case_data.get('id', ''),
                'Description': base_test_case_data.get('description', ''),
                'Precondition': base_test_case_data.get('precondition', ''),
                'Test Step #': step.step_number,
                'Test Step Description': step.description,
                'Test Step Expected Result': step.expected_result,
                'Action': step.action,
                'Notes': step.notes
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # Ensure column order matches QTEST + our additions
        column_order = [
            'Name', 'Id', 'Description', 'Precondition', 
            'Test Step #', 'Test Step Description', 'Test Step Expected Result',
            'Action', 'Notes'
        ]
        
        return df[column_order]
    
    def _apply_excel_formatting(self, worksheet):
        """Apply formatting to Excel worksheet for better readability"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Action column color coding
        action_colors = {
            'ADD': PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),      # Light green
            'MODIFY': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),   # Light yellow
            'DELETE': PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid")    # Light red
        }
        
        # Find Action column (should be column H)
        action_col_idx = None
        for idx, cell in enumerate(worksheet[1], 1):
            if cell.value == 'Action':
                action_col_idx = idx
                break
        
        if action_col_idx:
            # Apply color coding to Action column
            for row in worksheet.iter_rows(min_row=2, min_col=action_col_idx, max_col=action_col_idx):
                cell = row[0]
                if cell.value in action_colors:
                    cell.fill = action_colors[cell.value]
                    cell.font = Font(bold=True)
        
        # Adjust column widths
        column_widths = {
            'A': 50,  # Name
            'B': 12,  # Id  
            'C': 60,  # Description
            'D': 40,  # Precondition
            'E': 12,  # Test Step #
            'F': 80,  # Test Step Description
            'G': 80,  # Test Step Expected Result
            'H': 12,  # Action
            'I': 60   # Notes
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # Enable text wrapping for long columns
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def export_with_original_test_data(self, generated_steps: List[GeneratedTestStep],
                                     original_qtest_file: str) -> str:
        """Export modifications alongside original test case data"""
        
        # Read original QTEST file to get test case metadata
        try:
            original_df = pd.read_excel(original_qtest_file, sheet_name=0)
            
            # Get the first row of original data for base test case info
            if len(original_df) > 0:
                first_row = original_df.iloc[0]
                base_test_case_data = {
                    'name': first_row.get('Name', ''),
                    'id': first_row.get('Id', ''),
                    'description': first_row.get('Description', ''),
                    'precondition': first_row.get('Precondition', '')
                }
            else:
                self.logger.warning("Original QTEST file appears to be empty")
                base_test_case_data = self._get_default_test_case_data()
        
        except Exception as e:
            self.logger.error(f"Error reading original QTEST file {original_qtest_file}: {e}")
            base_test_case_data = self._get_default_test_case_data()
        
        # Export with the base test case data
        return self.export_to_excel(generated_steps, base_test_case_data, "from_qtest")
    
    def _get_default_test_case_data(self) -> Dict[str, Any]:
        """Get default test case data when original file is not available"""
        return {
            'name': 'Test Case Name (Update Required)',
            'id': 'TC-XXXX',
            'description': 'Test case description (Update Required)', 
            'precondition': 'Preconditions (Update Required)'
        }
    
    def export_summary_report(self, generated_steps: List[GeneratedTestStep],
                            generation_summary: Dict[str, Any]) -> str:
        """Export a summary report of all generated modifications"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_modifications_summary_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        
        # Create summary sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            
            # Sheet 1: Summary statistics
            summary_data = []
            summary_data.append(['Generation Date', generation_summary.get('generation_timestamp', '')])
            summary_data.append(['Total Steps Generated', generation_summary.get('total_steps_generated', 0)])
            summary_data.append(['', ''])
            summary_data.append(['Action Breakdown', ''])
            
            for action, count in generation_summary.get('action_breakdown', {}).items():
                summary_data.append([f'  {action}', count])
            
            summary_data.append(['', ''])
            summary_data.append(['Step Types', ''])
            
            for step_type, count in generation_summary.get('step_types', {}).items():
                summary_data.append([f'  {step_type.replace("_", " ").title()}', count])
            
            summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 2: Detailed steps
            if generated_steps:
                base_data = self._get_default_test_case_data()
                detailed_df = self._create_qtest_dataframe(generated_steps, base_data)
                detailed_df.to_excel(writer, sheet_name='Generated Steps', index=False)
                self._apply_excel_formatting(writer.book['Generated Steps'])
        
        self.logger.info(f"Summary report exported to {output_path}")
        return str(output_path)
    
    def create_instructions_sheet(self, output_path: str):
        """Add an instructions sheet to help QA team understand the modifications"""
        
        instructions = [
            ['STTM Test Modifications - Instructions', ''],
            ['', ''],
            ['This file contains test step modifications based on STTM changes.', ''],
            ['Each row represents an action to take on your test cases.', ''],
            ['', ''],
            ['Action Types:', ''],
            ['ADD', 'Create new test step with the provided description and expected result'],
            ['MODIFY', 'Update existing test step with new description/expected result'],
            ['DELETE', 'Remove existing test step (it references deleted STTM fields)'],
            ['', ''],
            ['Instructions:', ''],
            ['1. Review all ADD actions and create new test steps', ''],
            ['2. Review all MODIFY actions and update existing steps', ''],
            ['3. Review all DELETE actions and remove obsolete steps', ''],
            ['4. Update test case execution after implementing changes', ''],
            ['5. Validate all changes before marking test case as updated', ''],
            ['', ''],
            ['Notes Column:', 'Contains additional context about why each change is needed'],
            ['', ''],
            ['Questions?', 'Contact the STTM Impact Analysis team for clarification']
        ]
        
        try:
            # Read existing file and add instructions sheet
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                instructions_df = pd.DataFrame(instructions, columns=['Instruction', 'Details'])
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
                
                # Format instructions sheet
                ws = writer.book['Instructions']
                ws.column_dimensions['A'].width = 60
                ws.column_dimensions['B'].width = 80
                
                # Make first row a title
                try:
                    from openpyxl.styles import Font
                    title_font = Font(bold=True, size=14)
                    ws['A1'].font = title_font
                except ImportError:
                    self.logger.warning("openpyxl.styles.Font not available, skipping title formatting")
                
        except Exception as e:
            self.logger.error(f"Error adding instructions sheet: {e}")
    
    def validate_export_format(self, excel_file: str) -> bool:
        """Validate that the exported Excel file matches QTEST format requirements"""
        
        try:
            df = pd.read_excel(excel_file)
            
            # Check required QTEST columns
            required_columns = [
                'Name', 'Id', 'Description', 'Precondition',
                'Test Step #', 'Test Step Description', 'Test Step Expected Result'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                self.logger.error(f"Missing required columns: {missing_columns}")
                return False
            
            # Check for our Action column
            if 'Action' not in df.columns:
                self.logger.error("Missing Action column")
                return False
            
            # Validate Action values
            valid_actions = {'ADD', 'MODIFY', 'DELETE'}
            invalid_actions = df[~df['Action'].isin(valid_actions)]['Action'].unique()
            if len(invalid_actions) > 0:
                self.logger.error(f"Invalid Action values found: {invalid_actions}")
                return False
            
            self.logger.info(f"Export format validation passed for {excel_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating export format: {e}")
            return False
    
    def copy_and_modify_original(self, generated_steps: List[GeneratedTestStep], 
                                original_qtest_file: str) -> str:
        """Copy original QTEST file and make modifications in-place"""
        
        import shutil
        from pathlib import Path
        
        if not generated_steps:
            self.logger.warning("No generated steps to apply to original file")
            return ""
        
        # Create timestamped copy of original file
        original_path = Path(original_qtest_file)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        copied_filename = f"modified_{timestamp}_{original_path.name}"
        copied_path = self.output_dir / copied_filename
        
        self.logger.info(f"Creating copy of original QTEST file: {copied_path}")
        shutil.copy(original_qtest_file, copied_path)
        
        try:
            # Read second sheet (first sheet is cover page)
            self.logger.info("Reading test case data from second sheet")
            df = pd.read_excel(copied_path, sheet_name=1)  # Second sheet (index 1)
            
            # Find last step number for new additions
            if 'Test Step #' in df.columns:
                max_step = df['Test Step #'].max()
                self.logger.info(f"Found maximum existing step number: {max_step}")
            else:
                self.logger.error("Could not find 'Test Step #' column in QTEST file")
                return str(copied_path)
            
            # Process modifications
            next_step_number = max_step + 1
            new_rows = []
            
            for step in generated_steps:
                if step.action == 'ADD':
                    # Create new row for ADD actions (append at end)
                    new_row = self._create_new_row_from_existing(df, step, next_step_number)
                    new_rows.append(new_row)
                    self.logger.info(f"Adding new step {next_step_number}: {step.description[:50]}...")
                    next_step_number += 1
                    
                elif step.action == 'MODIFY':
                    # Find and update existing step
                    step_mask = df['Test Step #'] == step.step_number
                    if step_mask.any():
                        df.loc[step_mask, 'Test Step Description'] = step.description
                        df.loc[step_mask, 'Test Step Expected Result'] = step.expected_result
                        self.logger.info(f"Modified existing step {step.step_number}")
                    else:
                        self.logger.warning(f"Could not find existing step {step.step_number} to modify")
                
                elif step.action == 'DELETE':
                    # For DELETE, add comment to existing step (don't actually delete)
                    step_mask = df['Test Step #'] == step.step_number
                    if step_mask.any():
                        # Add note about deletion instead of actual deletion
                        current_desc = df.loc[step_mask, 'Test Step Description'].iloc[0]
                        df.loc[step_mask, 'Test Step Description'] = f"{current_desc} [NOTE: {step.notes}]"
                        self.logger.info(f"Added deletion note to step {step.step_number}")
                    else:
                        self.logger.warning(f"Could not find existing step {step.step_number} for deletion note")
            
            # Add new rows to dataframe
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                df = pd.concat([df, new_df], ignore_index=True)
                self.logger.info(f"Added {len(new_rows)} new test steps")
            
            # Write back to Excel file
            self.logger.info(f"Writing modified data back to {copied_path}")
            with pd.ExcelWriter(copied_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=writer.book.sheetnames[1], index=False)
            
            # Skip validation for in-place modifications since we preserve original structure
            self.logger.info(f"Successfully created modified QTEST file: {copied_path}")
            
            return str(copied_path)
            
        except Exception as e:
            self.logger.error(f"Error modifying original QTEST file: {e}")
            return str(copied_path)
    
    def _create_new_row_from_existing(self, df: pd.DataFrame, step: GeneratedTestStep, 
                                    step_number: int) -> Dict[str, Any]:
        """Create new row based on existing dataframe structure"""
        
        # Get first row as template for structure
        if len(df) > 0:
            first_row = df.iloc[0].to_dict()
            
            # Update with new step data
            new_row = {
                'Name': first_row.get('Name', ''),
                'Id': first_row.get('Id', ''),  
                'Description': first_row.get('Description', ''),
                'Precondition': first_row.get('Precondition', ''),
                'Test Step #': step_number,
                'Test Step Description': step.description,
                'Test Step Expected Result': step.expected_result
            }
            
            # Add any additional columns that exist in the original
            for col in df.columns:
                if col not in new_row:
                    new_row[col] = ''
                    
            return new_row
        else:
            # Fallback if no existing data
            return {
                'Name': 'Test Case Name',
                'Id': 'TC-XXXX',
                'Description': 'Test case description', 
                'Precondition': 'Preconditions',
                'Test Step #': step_number,
                'Test Step Description': step.description,
                'Test Step Expected Result': step.expected_result
            }