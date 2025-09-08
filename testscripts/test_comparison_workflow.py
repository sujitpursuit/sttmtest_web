"""
Test Comparison-Based Workflow
Tests the complete end-to-end workflow for comparison-based impact analysis and test step generation
"""

import requests
import json
import sys
from pathlib import Path
import time

# API base URL
API_BASE_URL = "http://127.0.0.1:8000"

# Test QTest file path (you'll need to provide a real Excel file)
QTEST_FILE_PATH = Path(__file__).parent.parent / "input_files" / "qtest" / "sample_qtest.xlsx"


def test_workflow():
    """Test the complete comparison-based workflow"""
    print("\n" + "="*60)
    print("COMPARISON-BASED WORKFLOW TEST")
    print("="*60)
    
    # Step 1: Get tracked files
    print("\nSTEP 1: Getting tracked files...")
    try:
        response = requests.get(f"{API_BASE_URL}/api/tracked-files", timeout=10)
        if response.status_code != 200:
            print(f"[FAIL] Failed to get tracked files: {response.status_code}")
            return False
            
        files_data = response.json()
        if not files_data.get('success') or not files_data.get('files'):
            print(f"[FAIL] No tracked files available")
            return False
            
        file_id = files_data['files'][0]['id']
        file_name = files_data['files'][0]['friendly_name']
        print(f"[PASS] Found tracked file: {file_name} (ID: {file_id})")
        
    except Exception as e:
        print(f"[FAIL] Error getting tracked files: {e}")
        return False
    
    # Step 2: Get comparisons for the file
    print(f"\nSTEP 2: Getting comparisons for file {file_id}...")
    try:
        response = requests.get(f"{API_BASE_URL}/api/tracked-files/{file_id}/comparisons", timeout=10)
        if response.status_code != 200:
            print(f"[FAIL] Failed to get comparisons: {response.status_code}")
            return False
            
        comp_data = response.json()
        if not comp_data.get('success') or not comp_data.get('comparisons'):
            print(f"[FAIL] No comparisons available")
            return False
            
        comparison = comp_data['comparisons'][0]  # Use first comparison
        comparison_id = comparison['comparison_id']
        comparison_title = comparison['comparison_title']
        print(f"[PASS] Found comparison: {comparison_title} (ID: {comparison_id})")
        print(f"        Changes: {comparison['total_changes']}")
        print(f"        JSON URL exists: {'Yes' if comparison.get('json_report_url') else 'No'}")
        
    except Exception as e:
        print(f"[FAIL] Error getting comparisons: {e}")
        return False
    
    # Step 3: Test Impact Analysis
    print(f"\nSTEP 3: Testing Impact Analysis with comparison {comparison_id}...")
    if not QTEST_FILE_PATH.exists():
        print(f"[WARNING] QTest file not found at {QTEST_FILE_PATH}")
        print(f"          Creating mock QTest file for testing...")
        # Create a simple mock Excel file
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Test Case ID'
            ws['B1'] = 'Test Case Name'
            ws['A2'] = 'TC001'
            ws['B2'] = 'Sample Test Case'
            QTEST_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
            wb.save(QTEST_FILE_PATH)
            print(f"[INFO] Created mock QTest file at {QTEST_FILE_PATH}")
        except ImportError:
            print(f"[FAIL] openpyxl not available and no QTest file found")
            return False
    
    try:
        with open(QTEST_FILE_PATH, 'rb') as f:
            files = {
                'qtest_file': ('sample_qtest.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            params = {
                'comparison_id': comparison_id
            }
            
            print(f"        Sending POST request to /api/analyze-impact-from-comparison...")
            response = requests.post(
                f"{API_BASE_URL}/api/analyze-impact-from-comparison",
                files=files,
                params=params,
                timeout=30
            )
            
        if response.status_code != 200:
            print(f"[FAIL] Impact analysis failed: {response.status_code}")
            print(f"        Response: {response.text[:500]}")
            return False
            
        impact_result = response.json()
        print(f"[PASS] Impact Analysis completed successfully!")
        
        # Check response structure
        if 'summary' in impact_result:
            summary = impact_result['summary']
            print(f"        Total STTM Changes: {summary.get('total_sttm_changes', 'N/A')}")
            print(f"        Total Test Cases: {summary.get('total_test_cases', 'N/A')}")
            print(f"        Critical Impacts: {summary.get('critical_impacts', 'N/A')}")
            print(f"        High Impacts: {summary.get('high_impacts', 'N/A')}")
        
        if 'comparison_info' in impact_result:
            comp_info = impact_result['comparison_info']
            print(f"        Comparison: {comp_info.get('comparison_title', 'N/A')}")
        
        # Check for HTML report links
        if 'reports' in impact_result:
            reports = impact_result['reports']
            print(f"        HTML Report: {'Available' if reports.get('html_file') else 'Not Available'}")
            print(f"        JSON Report: {'Available' if reports.get('json_file') else 'Not Available'}")
        
        print(f"        Response keys: {list(impact_result.keys())}")
        
    except Exception as e:
        print(f"[FAIL] Error during impact analysis: {e}")
        return False
    
    # Step 4: Test Test Step Generation
    print(f"\nSTEP 4: Testing Test Step Generation...")
    try:
        with open(QTEST_FILE_PATH, 'rb') as f:
            files = {
                'qtest_file': ('sample_qtest.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            params = {
                'comparison_id': comparison_id,
                'generation_mode': 'delta'
            }
            
            print(f"        Sending POST request to /api/generate/test-steps-from-comparison...")
            response = requests.post(
                f"{API_BASE_URL}/api/generate/test-steps-from-comparison",
                files=files,
                params=params,
                timeout=30
            )
            
        if response.status_code != 200:
            print(f"[FAIL] Test step generation failed: {response.status_code}")
            print(f"        Response: {response.text[:500]}")
            return False
            
        generation_result = response.json()
        print(f"[PASS] Test Step Generation completed successfully!")
        
        # Check response structure
        if 'summary' in generation_result:
            summary = generation_result['summary']
            print(f"        Generation Mode: {summary.get('generation_mode', 'N/A')}")
            print(f"        Total Test Cases Generated: {summary.get('total_test_cases_generated', 'N/A')}")
            print(f"        Files Created: {summary.get('files_created', 'N/A')}")
        
        if 'comparison_info' in generation_result:
            comp_info = generation_result['comparison_info']
            print(f"        Comparison: {comp_info.get('comparison_title', 'N/A')}")
        
        # Check for generated files
        if 'files' in generation_result:
            files_info = generation_result['files']
            print(f"        Generated Files:")
            for file_info in files_info:
                print(f"          - {file_info.get('filename', 'N/A')} ({file_info.get('type', 'N/A')})")
        
        print(f"        Response keys: {list(generation_result.keys())}")
        
    except Exception as e:
        print(f"[FAIL] Error during test step generation: {e}")
        return False
    
    return True


def main():
    """Run the complete workflow test"""
    print("Testing Comparison-Based Workflow Endpoints...")
    
    # Test API health first
    try:
        response = requests.get(f"{API_BASE_URL}/api/health", timeout=5)
        if response.status_code != 200:
            print(f"[FAIL] API not healthy: {response.status_code}")
            return False
        print(f"[PASS] API is healthy")
    except Exception as e:
        print(f"[FAIL] Cannot connect to API: {e}")
        print(f"       Make sure the API server is running on {API_BASE_URL}")
        return False
    
    # Run workflow test
    success = test_workflow()
    
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    
    if success:
        print("*** WORKFLOW TEST PASSED! ***")
        print("Both Impact Analysis and Test Step Generation work correctly")
        print("with the new comparison-based endpoints.")
    else:
        print("*** WORKFLOW TEST FAILED ***")
        print("Check the errors above and ensure:")
        print("1. API server is running")
        print("2. Database connection is working")
        print("3. Azure Blob Storage access is configured")
        print("4. QTest file is available")
    
    return success


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)