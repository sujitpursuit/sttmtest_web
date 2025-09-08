"""
Simple Azure Blob Storage Test
Tests Azure connection and upload without Unicode characters
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def main():
    """Simple test of Azure Blob Storage"""
    print("AZURE BLOB STORAGE SIMPLE TEST")
    print("=" * 50)
    
    try:
        # Test imports
        from azure.storage.blob import BlobServiceClient
        logger.info("[PASS] Azure libraries imported")
    except ImportError as e:
        logger.error(f"[FAIL] Azure import failed: {e}")
        return False
    
    # Test environment
    azure_conn = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
    if not azure_conn:
        logger.error("[FAIL] AZURE_STORAGE_CONNECTION_STRING not set")
        return False
    
    logger.info("[PASS] Connection string configured")
    
    try:
        # Test connection
        logger.info("Testing Azure connection...")
        blob_service_client = BlobServiceClient.from_connection_string(azure_conn)
        
        # List containers to test connection
        containers = list(blob_service_client.list_containers())
        logger.info(f"[PASS] Connected! Found {len(containers)} containers")
        
        # Test container
        container_name = "qtest-files"
        container_client = blob_service_client.get_container_client(container_name)
        
        if not container_client.exists():
            logger.info(f"Creating container: {container_name}")
            container_client.create_container()
            logger.info("[PASS] Container created")
        else:
            logger.info("[PASS] Container exists")
        
        # Create test file
        test_file = Path(__file__).parent / "azure_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        ws['B1'] = 'Data'
        wb.save(test_file)
        
        # Test upload
        with open(test_file, 'rb') as f:
            content = f.read()
        
        blob_name = f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        blob_client = container_client.get_blob_client(blob_name)
        
        # Test with ContentSettings (same fix as main service)
        from azure.storage.blob import ContentSettings
        content_settings = ContentSettings(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        logger.info(f"Uploading blob: {blob_name}")
        blob_client.upload_blob(content, overwrite=True, content_settings=content_settings)
        
        blob_url = blob_client.url
        logger.info(f"[PASS] Upload successful!")
        logger.info(f"Blob URL: {blob_url}")
        
        # Cleanup
        blob_client.delete_blob()
        test_file.unlink()
        logger.info("[PASS] Cleanup completed")
        
        print("=" * 50)
        print("ALL TESTS PASSED!")
        print("Azure Blob Storage is working correctly")
        return True
        
    except Exception as e:
        logger.error(f"[FAIL] Azure test failed: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        print("=" * 50)
        print("TEST FAILED!")
        print(f"Error: {e}")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)