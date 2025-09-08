"""
Test Azure Blob Storage Upload
Tests the Azure Blob Storage connection and upload functionality
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def test_azure_imports():
    """Test if Azure libraries are properly installed"""
    try:
        from azure.storage.blob import BlobServiceClient
        from azure.core.exceptions import AzureError
        logger.info("[PASS] Azure Storage libraries imported successfully")
        return True
    except ImportError as e:
        logger.error(f"[FAIL] Failed to import Azure libraries: {e}")
        logger.error("   Run: pip install azure-storage-blob")
        return False

def test_environment_variables():
    """Test if required environment variables are set"""
    azure_conn = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
    sql_conn = os.getenv('AZURE_SQL_CONNECTION_STRING')
    
    if not azure_conn:
        logger.error("[FAIL] AZURE_STORAGE_CONNECTION_STRING not set in environment")
        return False, None, None
    
    if not sql_conn:
        logger.error("[FAIL] AZURE_SQL_CONNECTION_STRING not set in environment")
        return False, azure_conn, None
        
    logger.info("[PASS] Both connection strings are configured")
    
    # Mask the connection strings for logging
    masked_azure = azure_conn[:20] + "..." + azure_conn[-10:] if len(azure_conn) > 30 else "***"
    masked_sql = sql_conn[:20] + "..." + sql_conn[-10:] if len(sql_conn) > 30 else "***"
    
    logger.info(f"   Azure Storage: {masked_azure}")
    logger.info(f"   Azure SQL: {masked_sql}")
    
    return True, azure_conn, sql_conn

def create_test_file():
    """Create a test QTest Excel file"""
    test_file_path = Path(__file__).parent / "test_azure_upload.xlsx"
    
    # Create a workbook with test data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Add headers
    headers = ['Test Case ID', 'Test Case Name', 'Description', 'Priority', 'Status']
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    # Add test data
    test_cases = [
        ['TC001', 'Login Functionality Test', 'Test user login with valid credentials', 'High', 'Active'],
        ['TC002', 'Dashboard Loading Test', 'Test dashboard loads correctly after login', 'Medium', 'Active'],
        ['TC003', 'User Profile Update', 'Test updating user profile information', 'Medium', 'Active'],
        ['TC004', 'Password Reset Test', 'Test password reset functionality', 'High', 'Active'],
        ['TC005', 'Logout Functionality', 'Test user logout process', 'Low', 'Active']
    ]
    
    for i, test_case in enumerate(test_cases, start=2):
        for j, value in enumerate(test_case, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # Add a second worksheet
    ws2 = wb.create_sheet("Test Steps")
    ws2['A1'] = 'Test Case ID'
    ws2['B1'] = 'Step Number'  
    ws2['C1'] = 'Step Description'
    ws2['D1'] = 'Expected Result'
    
    wb.save(test_file_path)
    logger.info(f"[PASS] Created test file: {test_file_path}")
    logger.info(f"   File size: {test_file_path.stat().st_size} bytes")
    
    return test_file_path

def test_blob_service_connection(connection_string):
    """Test connection to Azure Blob Service"""
    try:
        from azure.storage.blob import BlobServiceClient
        from azure.core.exceptions import AzureError
        
        logger.info("🔗 Testing Azure Blob Service connection...")
        
        # Create blob service client
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        
        # Test connection by listing containers (this will fail if connection is invalid)
        containers = list(blob_service_client.list_containers())
        logger.info(f"✅ Connection successful! Found {len(containers)} containers")
        
        for container in containers[:5]:  # Show first 5 containers
            logger.info(f"   📁 Container: {container.name}")
        
        return True, blob_service_client
        
    except Exception as e:
        logger.error(f"❌ Failed to connect to Azure Blob Storage: {e}")
        logger.error(f"   Error type: {type(e).__name__}")
        return False, None

def test_container_operations(blob_service_client, container_name="qtest-files"):
    """Test container creation and operations"""
    try:
        logger.info(f"📦 Testing container operations for '{container_name}'...")
        
        # Get container client
        container_client = blob_service_client.get_container_client(container_name)
        
        # Check if container exists
        exists = container_client.exists()
        logger.info(f"   Container exists: {exists}")
        
        if not exists:
            logger.info(f"   Creating container '{container_name}'...")
            container_client.create_container()
            logger.info(f"✅ Container '{container_name}' created successfully")
        else:
            logger.info(f"✅ Container '{container_name}' already exists")
        
        # List blobs in container
        blobs = list(container_client.list_blobs())
        logger.info(f"   Found {len(blobs)} blobs in container")
        
        for blob in blobs[:3]:  # Show first 3 blobs
            logger.info(f"   📄 Blob: {blob.name} ({blob.size} bytes)")
        
        return True, container_client
        
    except Exception as e:
        logger.error(f"❌ Container operations failed: {e}")
        logger.error(f"   Error type: {type(e).__name__}")
        return False, None

def test_blob_upload(container_client, test_file_path, comparison_id=123):
    """Test blob upload functionality"""
    try:
        logger.info(f"📤 Testing blob upload...")
        
        # Generate blob name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        blob_name = f"comparison_{comparison_id}/qtest_{timestamp}_{test_file_path.name}"
        
        logger.info(f"   Blob name: {blob_name}")
        
        # Read file content
        with open(test_file_path, 'rb') as f:
            file_content = f.read()
        
        logger.info(f"   File size: {len(file_content)} bytes")
        
        # Get blob client
        blob_client = container_client.get_blob_client(blob_name)
        
        # Upload blob
        logger.info("   Uploading to Azure Blob Storage...")
        blob_client.upload_blob(
            file_content,
            overwrite=True,
            content_settings={
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
        
        # Get blob URL
        blob_url = blob_client.url
        logger.info(f"✅ Upload successful!")
        logger.info(f"   Blob URL: {blob_url}")
        
        # Verify upload by getting blob properties
        properties = blob_client.get_blob_properties()
        logger.info(f"   Uploaded size: {properties.size} bytes")
        logger.info(f"   Content type: {properties.content_settings.content_type}")
        logger.info(f"   Last modified: {properties.last_modified}")
        
        return True, blob_url, blob_name
        
    except Exception as e:
        logger.error(f"❌ Blob upload failed: {e}")
        logger.error(f"   Error type: {type(e).__name__}")
        logger.error(f"   Error details: {str(e)}")
        return False, None, None

def test_blob_download(container_client, blob_name):
    """Test blob download functionality"""
    try:
        logger.info(f"📥 Testing blob download...")
        
        blob_client = container_client.get_blob_client(blob_name)
        
        # Download blob
        blob_data = blob_client.download_blob()
        content = blob_data.readall()
        
        logger.info(f"✅ Download successful!")
        logger.info(f"   Downloaded size: {len(content)} bytes")
        
        return True, content
        
    except Exception as e:
        logger.error(f"❌ Blob download failed: {e}")
        return False, None

def cleanup_test_blob(container_client, blob_name):
    """Clean up test blob"""
    try:
        logger.info(f"🧹 Cleaning up test blob...")
        
        blob_client = container_client.get_blob_client(blob_name)
        blob_client.delete_blob()
        
        logger.info(f"✅ Test blob deleted: {blob_name}")
        
    except Exception as e:
        logger.warning(f"⚠️ Failed to cleanup blob: {e}")

def cleanup_test_file(test_file_path):
    """Clean up test file"""
    try:
        if test_file_path.exists():
            test_file_path.unlink()
            logger.info(f"🧹 Cleaned up test file: {test_file_path}")
    except Exception as e:
        logger.warning(f"⚠️ Could not clean up test file: {e}")

def main():
    """Run complete Azure Blob Storage test suite"""
    print("AZURE BLOB STORAGE TEST SUITE")
    print("=" * 60)
    
    success = True
    
    # Test 1: Azure imports
    if not test_azure_imports():
        return False
    
    # Test 2: Environment variables
    env_ok, azure_conn, sql_conn = test_environment_variables()
    if not env_ok:
        logger.error("💡 To fix: Add AZURE_STORAGE_CONNECTION_STRING to your .env file")
        return False
    
    # Test 3: Create test file
    test_file_path = create_test_file()
    
    container_client = None
    blob_name = None
    
    try:
        # Test 4: Blob service connection
        conn_ok, blob_service_client = test_blob_service_connection(azure_conn)
        if not conn_ok:
            success = False
            return False
        
        # Test 5: Container operations
        container_ok, container_client = test_container_operations(blob_service_client)
        if not container_ok:
            success = False
            return False
        
        # Test 6: Blob upload
        upload_ok, blob_url, blob_name = test_blob_upload(container_client, test_file_path)
        if not upload_ok:
            success = False
            return False
        
        # Test 7: Blob download (verification)
        download_ok, content = test_blob_download(container_client, blob_name)
        if not download_ok:
            success = False
        
        # Verify file integrity
        if download_ok:
            with open(test_file_path, 'rb') as f:
                original_content = f.read()
            
            if len(content) == len(original_content) and content == original_content:
                logger.info("✅ File integrity verified - upload/download successful!")
            else:
                logger.error("❌ File integrity check failed")
                success = False
        
    finally:
        # Cleanup
        if container_client and blob_name:
            cleanup_test_blob(container_client, blob_name)
        cleanup_test_file(test_file_path)
    
    print("\n" + "=" * 60)
    print("🏁 TEST RESULTS")
    print("=" * 60)
    
    if success:
        print("🎉 ALL TESTS PASSED!")
        print("✅ Azure Blob Storage is working correctly")
        print("✅ Connection string is valid")
        print("✅ Container operations work")
        print("✅ File upload/download works")
        print("✅ File integrity is maintained")
        print("\nYour Azure configuration is ready for QTest uploads!")
    else:
        print("❌ SOME TESTS FAILED")
        print("Please check the errors above and verify:")
        print("- Azure Storage connection string is correct")
        print("- Azure Storage account has proper permissions")
        print("- Network connectivity to Azure is available")
    
    return success

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)