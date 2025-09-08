"""
Test Chrome ERR_UPLOAD_FILE_CHANGED Fix
Tests the new two-step process: upload file first, then analyze
"""

import requests
import json
import sys
from pathlib import Path

# API base URL
API_BASE_URL = "http://127.0.0.1:8000"

# Test QTest file path
QTEST_FILE_PATH = Path(__file__).parent.parent / "input_files" / "qtest" / "sample_qtest.xlsx"


def test_two_step_process():
    """Test the new two-step process that should work in Chrome"""
    print("\n" + "="*60)
    print("CHROME FIX TEST - Two-Step Process")
    print("="*60)
    
    # Step 1: Get a comparison to test with
    print("\nSTEP 1: Getting comparison for testing...")
    try:
        # Get tracked files
        response = requests.get(f"{API_BASE_URL}/api/tracked-files", timeout=10)
        if response.status_code != 200:
            print(f"[FAIL] Failed to get tracked files: {response.status_code}")
            return False
            
        files_data = response.json()
        if not files_data.get('success') or not files_data.get('files'):
            print(f"[FAIL] No tracked files available")
            return False
            
        file_id = files_data['files'][0]['id']
        print(f"[PASS] Found tracked file (ID: {file_id})")
        
        # Get comparisons
        response = requests.get(f"{API_BASE_URL}/api/tracked-files/{file_id}/comparisons", timeout=10)
        if response.status_code != 200:
            print(f"[FAIL] Failed to get comparisons: {response.status_code}")
            return False
            
        comp_data = response.json()
        if not comp_data.get('success') or not comp_data.get('comparisons'):
            print(f"[FAIL] No comparisons available")
            return False
            
        comparison_id = comp_data['comparisons'][0]['comparison_id']
        comparison_title = comp_data['comparisons'][0]['comparison_title']
        print(f"[PASS] Found comparison: {comparison_title} (ID: {comparison_id})")
        
    except Exception as e:
        print(f"[FAIL] Error getting comparison: {e}")
        return False
    
    # Step 2: Test file upload (new separate endpoint)
    print(f"\nSTEP 2: Testing QTest file upload...")
    if not QTEST_FILE_PATH.exists():
        print(f"[WARNING] QTest file not found at {QTEST_FILE_PATH}")
        print(f"          Creating mock QTest file for testing...")
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Test Case ID'
            ws['B1'] = 'Test Case Name'
            ws['A2'] = 'TC001'
            ws['B2'] = 'Sample Test Case'
            QTEST_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
            wb.save(QTEST_FILE_PATH)
            print(f"[INFO] Created mock QTest file at {QTEST_FILE_PATH}")
        except ImportError:
            print(f"[FAIL] openpyxl not available and no QTest file found")
            return False
    
    try:
        with open(QTEST_FILE_PATH, 'rb') as f:
            files = {
                'file': ('sample_qtest.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            print(f"        Uploading file: {QTEST_FILE_PATH.name}")
            response = requests.post(
                f"{API_BASE_URL}/api/upload/qtest",
                files=files,
                timeout=30
            )
            
        if response.status_code != 200:
            print(f"[FAIL] File upload failed: {response.status_code}")
            print(f"        Response: {response.text[:500]}")
            return False
            
        upload_result = response.json()
        print(f"[PASS] File uploaded successfully!")
        print(f"        Filename: {upload_result['filename']}")
        print(f"        File size: {upload_result['file_size']} bytes")
        
    except Exception as e:
        print(f"[FAIL] Error during file upload: {e}")
        return False
    
    # Step 3: Test analysis with filename (no file upload)
    print(f"\nSTEP 3: Testing impact analysis with filename...")
    try:
        params = {
            'comparison_id': comparison_id,
            'qtest_file': upload_result['filename']
        }
        
        print(f"        Sending analysis request with params: {params}")
        response = requests.post(
            f"{API_BASE_URL}/api/analyze-impact-from-comparison",
            params=params,
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"[FAIL] Impact analysis failed: {response.status_code}")
            print(f"        Response: {response.text[:500]}")
            return False
            
        analysis_result = response.json()
        print(f"[PASS] Impact Analysis completed successfully!")
        
        # Check response structure
        if 'summary' in analysis_result:
            summary = analysis_result['summary']
            print(f"        Total STTM Changes: {summary.get('total_sttm_changes', 'N/A')}")
            print(f"        Total Test Cases: {summary.get('total_test_cases', 'N/A')}")
            print(f"        Critical Impacts: {summary.get('critical_impacts', 'N/A')}")
        
    except Exception as e:
        print(f"[FAIL] Error during impact analysis: {e}")
        return False
    
    # Step 4: Test step generation with filename
    print(f"\nSTEP 4: Testing test step generation with filename...")
    try:
        params = {
            'comparison_id': comparison_id,
            'generation_mode': 'delta',
            'qtest_file': upload_result['filename']
        }
        
        print(f"        Sending generation request with params: {params}")
        response = requests.post(
            f"{API_BASE_URL}/api/generate/test-steps-from-comparison",
            params=params,
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"[FAIL] Test step generation failed: {response.status_code}")
            print(f"        Response: {response.text[:500]}")
            return False
            
        generation_result = response.json()
        print(f"[PASS] Test Step Generation completed successfully!")
        
        # Check response structure
        if 'summary' in generation_result:
            summary = generation_result['summary']
            print(f"        Generation Mode: {summary.get('generation_mode', 'N/A')}")
            print(f"        Files Created: {summary.get('files_created', 'N/A')}")
        
    except Exception as e:
        print(f"[FAIL] Error during test step generation: {e}")
        return False
    
    return True


def main():
    """Run the Chrome fix test"""
    print("Testing Chrome ERR_UPLOAD_FILE_CHANGED Fix...")
    
    # Test API health first
    try:
        response = requests.get(f"{API_BASE_URL}/api/health", timeout=5)
        if response.status_code != 200:
            print(f"[FAIL] API not healthy: {response.status_code}")
            return False
        print(f"[PASS] API is healthy")
    except Exception as e:
        print(f"[FAIL] Cannot connect to API: {e}")
        print(f"       Make sure the API server is running on {API_BASE_URL}")
        return False
    
    # Run two-step process test
    success = test_two_step_process()
    
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    
    if success:
        print("*** TWO-STEP PROCESS TEST PASSED! ***")
        print("The new implementation should work in Chrome:")
        print("1. File upload is separated from analysis")
        print("2. No file object modifications during analysis")
        print("3. Uses the same pattern as the working original endpoints")
    else:
        print("*** TWO-STEP PROCESS TEST FAILED ***")
        print("Check the errors above and ensure:")
        print("1. API server is running")
        print("2. Database connection is working")
        print("3. Azure Blob Storage access is configured")
    
    return success


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)