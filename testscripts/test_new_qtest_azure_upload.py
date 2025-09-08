"""
Test Script for NEW QTest Azure Upload Functionality
Tests the complete workflow: validation, Azure Blob upload, database update
"""

import requests
import json
import sys
import os
from pathlib import Path
import openpyxl
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# API base URL
API_BASE_URL = "http://127.0.0.1:8000"

def create_test_qtest_file():
    """Create a simple test QTest Excel file"""
    test_file_path = Path(__file__).parent / "test_qtest_azure.xlsx"
    
    # Create a workbook with sample test data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Add headers
    ws['A1'] = 'Test Case ID'
    ws['B1'] = 'Test Case Name'
    ws['C1'] = 'Description'
    ws['D1'] = 'Steps'
    
    # Add sample test cases
    test_cases = [
        ['TC001', 'Login Test', 'Test user login functionality', 'Step 1: Open app\nStep 2: Enter credentials\nStep 3: Click login'],
        ['TC002', 'Dashboard Test', 'Test dashboard loading', 'Step 1: Login\nStep 2: Navigate to dashboard\nStep 3: Verify data'],
        ['TC003', 'Logout Test', 'Test user logout', 'Step 1: Click logout\nStep 2: Verify redirect'],
    ]
    
    for i, test_case in enumerate(test_cases, start=2):
        for j, value in enumerate(test_case):
            ws.cell(row=i, column=j+1, value=value)
    
    wb.save(test_file_path)
    print(f"Created test QTest file: {test_file_path}")
    return test_file_path

def test_api_health():
    """Test if the API is running and healthy"""
    try:
        response = requests.get(f"{API_BASE_URL}/api/health", timeout=5)
        if response.status_code == 200:
            print("✅ API is healthy and running")
            return True
        else:
            print(f"❌ API health check failed: {response.status_code}")
            return False
    except Exception as e:
        print(f"❌ Cannot connect to API: {e}")
        print(f"   Make sure the API server is running on {API_BASE_URL}")
        return False

def test_version_info():
    """Test version endpoint and confirm v1.0.3"""
    try:
        response = requests.get(f"{API_BASE_URL}/api/version", timeout=5)
        if response.status_code == 200:
            version_data = response.json()
            version = version_data.get('version_info', {}).get('version')
            if version == '1.0.3':
                print(f"✅ Version confirmed: {version}")
                return True
            else:
                print(f"⚠️ Expected v1.0.3, got v{version}")
                return True  # Still ok to continue
        else:
            print(f"❌ Version check failed: {response.status_code}")
            return False
    except Exception as e:
        print(f"❌ Version check error: {e}")
        return False

def get_test_comparison():
    """Get a comparison ID for testing"""
    try:
        # Get tracked files
        response = requests.get(f"{API_BASE_URL}/api/tracked-files", timeout=10)
        if response.status_code != 200:
            print(f"❌ Failed to get tracked files: {response.status_code}")
            return None
            
        files_data = response.json()
        if not files_data.get('success') or not files_data.get('files'):
            print("❌ No tracked files available")
            return None
            
        file_id = files_data['files'][0]['id']
        file_name = files_data['files'][0]['friendly_name']
        print(f"✅ Found tracked file: {file_name} (ID: {file_id})")
        
        # Get comparisons
        response = requests.get(f"{API_BASE_URL}/api/tracked-files/{file_id}/comparisons", timeout=10)
        if response.status_code != 200:
            print(f"❌ Failed to get comparisons: {response.status_code}")
            return None
            
        comp_data = response.json()
        if not comp_data.get('success') or not comp_data.get('comparisons'):
            print("❌ No comparisons available")
            return None
            
        comparison = comp_data['comparisons'][0]
        comparison_id = comparison['comparison_id']
        comparison_title = comparison['comparison_title']
        print(f"✅ Found comparison: {comparison_title} (ID: {comparison_id})")
        
        return comparison_id
        
    except Exception as e:
        print(f"❌ Error getting test comparison: {e}")
        return None

def test_new_qtest_upload(comparison_id, test_file_path):
    """Test the NEW QTest Azure upload endpoint"""
    print(f"\n🧪 Testing NEW QTest Azure Upload for comparison {comparison_id}")
    print("=" * 60)
    
    try:
        # Upload the test file
        with open(test_file_path, 'rb') as f:
            files = {
                'file': ('test_qtest_azure.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            print(f"📤 Uploading QTest file to Azure...")
            response = requests.post(
                f"{API_BASE_URL}/api/qtest/upload-validate/{comparison_id}",
                files=files,
                timeout=30
            )
            
        print(f"📊 Response Status: {response.status_code}")
        
        if response.status_code != 200:
            print(f"❌ Upload failed: {response.status_code}")
            print(f"   Response: {response.text[:500]}")
            return False
            
        result = response.json()
        print(f"📄 Response: {json.dumps(result, indent=2)}")
        
        # Verify the response structure
        if result.get('success'):
            print("✅ Upload marked as successful")
            
            validation = result.get('validation', {})
            if validation.get('valid'):
                print("✅ Validation passed")
                print(f"   📊 Worksheets: {validation.get('worksheets', 0)}")
                print(f"   📝 Test Cases: {validation.get('test_cases', 0)}")
                print(f"   📁 File Size: {validation.get('file_size', 0)} bytes")
                
                if validation.get('warnings'):
                    print(f"   ⚠️ Warnings: {', '.join(validation['warnings'])}")
            else:
                print("❌ Validation failed")
                print(f"   Errors: {validation.get('errors', [])}")
                return False
            
            blob_url = result.get('blob_url')
            if blob_url:
                print(f"✅ Azure Blob URL: {blob_url}")
            else:
                print("❌ No blob URL returned")
                return False
                
            if result.get('comparison_updated'):
                print("✅ Database updated successfully")
            else:
                print("❌ Database update failed")
                return False
                
            return True
        else:
            print(f"❌ Upload failed: {result.get('error', 'Unknown error')}")
            return False
            
    except Exception as e:
        print(f"❌ Upload test error: {e}")
        return False

def cleanup_test_file(test_file_path):
    """Clean up the test file"""
    try:
        if test_file_path.exists():
            test_file_path.unlink()
            print(f"🧹 Cleaned up test file: {test_file_path}")
    except Exception as e:
        print(f"⚠️ Could not clean up test file: {e}")

def main():
    """Run the complete test suite"""
    print("🧪 TESTING NEW QTEST AZURE UPLOAD FUNCTIONALITY")
    print("=" * 60)
    print(f"Target API: {API_BASE_URL}")
    print(f"Expected Version: 1.0.3")
    print("=" * 60)
    
    # Test 1: API Health
    if not test_api_health():
        return False
    
    # Test 2: Version Check
    if not test_version_info():
        return False
    
    # Test 3: Get test comparison
    comparison_id = get_test_comparison()
    if not comparison_id:
        return False
    
    # Test 4: Create test QTest file
    test_file_path = create_test_qtest_file()
    
    try:
        # Test 5: Test the new upload functionality
        success = test_new_qtest_upload(comparison_id, test_file_path)
        
        print("\n" + "=" * 60)
        print("🏁 TEST SUMMARY")
        print("=" * 60)
        
        if success:
            print("🎉 ALL TESTS PASSED!")
            print("✅ NEW QTest Azure upload functionality is working correctly")
            print("✅ Validation works")
            print("✅ Azure Blob Storage upload works")
            print("✅ Database update works")
            print("\nThe new implementation should resolve the Chrome upload issues!")
        else:
            print("❌ TESTS FAILED")
            print("Please check the errors above and ensure:")
            print("- Azure Storage connection string is configured")
            print("- Database is accessible")
            print("- All dependencies are installed")
            
        return success
        
    finally:
        # Cleanup
        cleanup_test_file(test_file_path)

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)