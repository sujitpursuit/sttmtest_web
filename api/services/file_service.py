"""
File Service - Manage input files for STTM Impact Analysis API
Handles file discovery, validation, and path resolution for input_files directory
"""

from pathlib import Path
from typing import List, Optional
import json
import logging


class FileService:
    """Service for managing STTM and QTEST input files"""
    
    def __init__(self):
        self.base_dir = Path(__file__).parent.parent.parent  # STTMQTEST_WEB root
        self.input_dir = self.base_dir / "input_files"
        self.sttm_dir = self.input_dir / "sttm"
        self.qtest_dir = self.input_dir / "qtest"
        self.reports_dir = self.base_dir / "reports"
        
        # Ensure directories exist
        self.input_dir.mkdir(exist_ok=True)
        self.sttm_dir.mkdir(exist_ok=True)
        self.qtest_dir.mkdir(exist_ok=True)
        self.reports_dir.mkdir(exist_ok=True)
        
        self.logger = logging.getLogger(__name__)
    
    def list_sttm_files(self) -> List[dict]:
        """Return list of available STTM JSON files with metadata"""
        files = []
        
        for file_path in self.sttm_dir.glob("*.json"):
            try:
                # Get basic file info
                stat = file_path.stat()
                file_info = {
                    "filename": file_path.name,
                    "size_bytes": stat.st_size,
                    "modified": stat.st_mtime,
                    "valid": True,
                    "error": None
                }
                
                # Try to validate JSON structure
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        json.load(f)
                except json.JSONDecodeError as e:
                    file_info["valid"] = False
                    file_info["error"] = f"Invalid JSON: {str(e)}"
                
                files.append(file_info)
                
            except Exception as e:
                self.logger.error(f"Error reading STTM file {file_path.name}: {e}")
                files.append({
                    "filename": file_path.name,
                    "valid": False,
                    "error": str(e)
                })
        
        return sorted(files, key=lambda x: x["filename"])
    
    def list_qtest_files(self) -> List[dict]:
        """Return list of available QTEST Excel files with metadata"""
        files = []
        
        for file_path in self.qtest_dir.glob("*.xlsx"):
            try:
                stat = file_path.stat()
                file_info = {
                    "filename": file_path.name,
                    "size_bytes": stat.st_size,
                    "modified": stat.st_mtime,
                    "valid": True,
                    "error": None
                }
                
                # Basic Excel validation - check if file can be opened
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    wb.close()
                except Exception as e:
                    file_info["valid"] = False
                    file_info["error"] = f"Cannot open Excel file: {str(e)}"
                
                files.append(file_info)
                
            except Exception as e:
                self.logger.error(f"Error reading QTEST file {file_path.name}: {e}")
                files.append({
                    "filename": file_path.name,
                    "valid": False,
                    "error": str(e)
                })
        
        return sorted(files, key=lambda x: x["filename"])
    
    def get_sttm_path(self, filename: str) -> str:
        """Get full path to STTM file, raise exception if not found"""
        if not filename.endswith('.json'):
            raise ValueError(f"STTM file must be JSON: {filename}")
        
        file_path = self.sttm_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"STTM file not found: {filename}")
        
        if not file_path.is_file():
            raise ValueError(f"Path is not a file: {filename}")
        
        return str(file_path)
    
    def get_qtest_path(self, filename: str) -> str:
        """Get full path to QTEST file, raise exception if not found"""
        if not filename.endswith('.xlsx'):
            raise ValueError(f"QTEST file must be Excel (.xlsx): {filename}")
        
        file_path = self.qtest_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"QTEST file not found: {filename}")
        
        if not file_path.is_file():
            raise ValueError(f"Path is not a file: {filename}")
        
        return str(file_path)
    
    def validate_sttm_file(self, filename: str) -> dict:
        """Validate STTM JSON file and return detailed validation results"""
        try:
            file_path = self.get_sttm_path(filename)
            
            # Try to parse JSON
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Basic structure validation
            validation_result = {
                "filename": filename,
                "valid": True,
                "file_size": Path(file_path).stat().st_size,
                "structure": {
                    "has_data": bool(data),
                    "keys_found": list(data.keys()) if isinstance(data, dict) else [],
                    "estimated_tabs": 0,
                    "estimated_changes": 0
                },
                "errors": [],
                "warnings": []
            }
            
            # Try to estimate content (basic heuristics)
            if isinstance(data, dict):
                # Look for tab-like structures
                for key, value in data.items():
                    if isinstance(value, (dict, list)):
                        validation_result["structure"]["estimated_tabs"] += 1
                        if isinstance(value, list):
                            validation_result["structure"]["estimated_changes"] += len(value)
                        elif isinstance(value, dict) and "changes" in str(value).lower():
                            validation_result["structure"]["estimated_changes"] += len(value)
            
            return validation_result
            
        except FileNotFoundError as e:
            return {
                "filename": filename,
                "valid": False,
                "errors": [str(e)],
                "warnings": []
            }
        except json.JSONDecodeError as e:
            return {
                "filename": filename, 
                "valid": False,
                "errors": [f"Invalid JSON format: {str(e)}"],
                "warnings": []
            }
        except Exception as e:
            return {
                "filename": filename,
                "valid": False,
                "errors": [f"Validation error: {str(e)}"],
                "warnings": []
            }
    
    def validate_qtest_file(self, filename: str) -> dict:
        """Validate QTEST Excel file and return detailed validation results"""
        try:
            file_path = self.get_qtest_path(filename)
            
            import openpyxl
            
            validation_result = {
                "filename": filename,
                "valid": True,
                "file_size": Path(file_path).stat().st_size,
                "structure": {
                    "worksheets": [],
                    "estimated_test_cases": 0,
                    "estimated_test_steps": 0
                },
                "errors": [],
                "warnings": []
            }
            
            # Open and analyze Excel structure
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet_info = {
                    "name": sheet_name,
                    "rows": 0,
                    "columns": 0
                }
                
                ws = wb[sheet_name]
                sheet_info["rows"] = ws.max_row or 0
                sheet_info["columns"] = ws.max_column or 0
                
                # Estimate test cases (rows with data)
                if sheet_info["rows"] > 1:  # Exclude header
                    validation_result["structure"]["estimated_test_cases"] += max(0, sheet_info["rows"] - 1)
                
                validation_result["structure"]["worksheets"].append(sheet_info)
            
            wb.close()
            
            # Add warnings for common issues
            if not validation_result["structure"]["worksheets"]:
                validation_result["warnings"].append("No worksheets found")
            elif validation_result["structure"]["estimated_test_cases"] == 0:
                validation_result["warnings"].append("No test cases detected")
            
            return validation_result
            
        except FileNotFoundError as e:
            return {
                "filename": filename,
                "valid": False,
                "errors": [str(e)],
                "warnings": []
            }
        except Exception as e:
            return {
                "filename": filename,
                "valid": False,
                "errors": [f"Excel validation error: {str(e)}"],
                "warnings": []
            }
    
    def get_file_info(self, file_type: str, filename: str) -> Optional[dict]:
        """Get detailed info about a specific file"""
        if file_type.lower() == "sttm":
            files = self.list_sttm_files()
        elif file_type.lower() == "qtest":
            files = self.list_qtest_files()
        else:
            return None
        
        for file_info in files:
            if file_info["filename"] == filename:
                return file_info
        
        return None