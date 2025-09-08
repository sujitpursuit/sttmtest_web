"""
QTest Azure Service - NEW implementation for QTest file handling
Handles validation, Azure Blob upload, and database updates
"""

import os
import logging
from datetime import datetime
from typing import Dict, Optional, Tuple
import openpyxl
from pathlib import Path
import pyodbc
from azure.storage.blob import BlobServiceClient, ContentSettings
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)


class QTestAzureService:
    """New service for QTest file operations with Azure Blob Storage"""
    
    def __init__(self):
        """Initialize Azure Blob and Database connections"""
        self.blob_connection_string = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
        self.container_name = "qtest-files"
        self.db_connection_string = os.getenv('AZURE_SQL_CONNECTION_STRING')
        
        if not self.blob_connection_string:
            logger.error("Azure Storage connection string not configured")
            raise ValueError("Azure Storage connection string is required")
        
        if not self.db_connection_string:
            logger.error("Azure SQL connection string not configured")
            raise ValueError("Azure SQL connection string is required")
        
        # Initialize blob service client
        self.blob_service_client = BlobServiceClient.from_connection_string(
            self.blob_connection_string
        )
        
        # Ensure container exists
        self._ensure_container_exists()
        
        logger.info(f"QTestAzureService initialized successfully")
        logger.info(f"Container: {self.container_name}")
    
    def _ensure_container_exists(self):
        """Create the qtest-files container if it doesn't exist"""
        try:
            container_client = self.blob_service_client.get_container_client(self.container_name)
            if not container_client.exists():
                container_client.create_container()
                logger.info(f"Created container: {self.container_name}")
        except Exception as e:
            logger.error(f"Error ensuring container exists: {e}")
    
    def validate_qtest_file(self, file_content: bytes, filename: str) -> Tuple[bool, Dict]:
        """
        Perform basic validation on QTest Excel file
        
        Returns:
            Tuple of (is_valid, validation_details)
        """
        validation_result = {
            "valid": False,
            "filename": filename,
            "file_size": len(file_content),
            "worksheets": 0,
            "test_cases": 0,
            "errors": [],
            "warnings": []
        }
        
        # Check file size (10MB limit)
        if len(file_content) > 10 * 1024 * 1024:
            validation_result["errors"].append("File size exceeds 10MB limit")
            return False, validation_result
        
        # Check file extension
        if not (filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')):
            validation_result["errors"].append("File must be Excel format (.xlsx or .xls)")
            return False, validation_result
        
        try:
            # Save temporarily to validate structure
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(file_content)
                tmp_path = tmp_file.name
            
            # Open and validate Excel structure
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            
            worksheet_count = len(wb.sheetnames)
            validation_result["worksheets"] = worksheet_count
            
            if worksheet_count == 0:
                validation_result["errors"].append("No worksheets found in Excel file")
                wb.close()
                os.unlink(tmp_path)
                return False, validation_result
            
            # Count test cases (assuming first sheet contains test cases)
            first_sheet = wb[wb.sheetnames[0]]
            row_count = first_sheet.max_row
            
            if row_count and row_count > 1:  # Exclude header row
                validation_result["test_cases"] = row_count - 1
            
            # Check for minimum content
            if validation_result["test_cases"] == 0:
                validation_result["warnings"].append("No test cases found in file")
            
            wb.close()
            os.unlink(tmp_path)
            
            # Mark as valid if no errors
            if not validation_result["errors"]:
                validation_result["valid"] = True
                logger.info(f"QTest validation successful: {filename}")
            
        except Exception as e:
            validation_result["errors"].append(f"Failed to parse Excel file: {str(e)}")
            logger.error(f"QTest validation error: {e}")
            return False, validation_result
        
        return validation_result["valid"], validation_result
    
    def upload_to_azure_blob(self, file_content: bytes, filename: str, comparison_id: int) -> Optional[str]:
        """
        Upload validated QTest file to Azure Blob Storage
        
        Returns:
            Blob URL if successful, None otherwise
        """
        try:
            # Create blob name with comparison context
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            blob_name = f"comparison_{comparison_id}/qtest_{timestamp}_{filename}"
            
            # Get blob client
            blob_client = self.blob_service_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            # Upload file with proper ContentSettings object
            content_settings = ContentSettings(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            blob_client.upload_blob(
                file_content,
                overwrite=True,
                content_settings=content_settings
            )
            
            # Get blob URL
            blob_url = blob_client.url
            logger.info(f"Successfully uploaded QTest to Azure: {blob_url}")
            
            return blob_url
            
        except Exception as e:
            logger.error(f"Failed to upload QTest to Azure: {e}")
            logger.error(f"Error type: {type(e).__name__}")
            logger.error(f"Blob name: {blob_name}")
            logger.error(f"File size: {len(file_content)} bytes")
            # Store the error for detailed reporting
            self.last_upload_error = f"{type(e).__name__}: {str(e)}"
            return None
    
    def update_comparison_qtest_url(self, comparison_id: int, blob_url: str) -> bool:
        """
        Update version_comparisons table with QTest blob URL
        
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Attempting to update comparison {comparison_id} with blob URL")
            
            with pyodbc.connect(self.db_connection_string) as conn:
                cursor = conn.cursor()
                
                # First check if the comparison exists
                check_query = "SELECT COUNT(*) FROM version_comparisons WHERE id = ?"
                cursor.execute(check_query, comparison_id)
                count = cursor.fetchone()[0]
                
                if count == 0:
                    logger.error(f"Comparison {comparison_id} not found in database")
                    self.last_db_error = f"Comparison ID {comparison_id} does not exist"
                    return False
                
                # Check if qtest_file column exists
                column_check = """
                SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = 'version_comparisons' 
                AND COLUMN_NAME = 'qtest_file'
                """
                cursor.execute(column_check)
                column_exists = cursor.fetchone()[0] > 0
                
                if not column_exists:
                    logger.error("qtest_file column does not exist in version_comparisons table")
                    self.last_db_error = "qtest_file column missing - run database migration"
                    return False
                
                # Update the record (without updated_at since it might not exist)
                query = """
                UPDATE version_comparisons 
                SET qtest_file = ?
                WHERE id = ?
                """
                
                logger.info(f"Executing update query with URL: {blob_url[:50]}...")
                cursor.execute(query, blob_url, comparison_id)
                conn.commit()
                
                if cursor.rowcount > 0:
                    logger.info(f"Successfully updated comparison {comparison_id} with QTest URL")
                    return True
                else:
                    logger.warning(f"Update executed but no rows affected for comparison {comparison_id}")
                    self.last_db_error = f"No rows affected - comparison {comparison_id} may not exist"
                    return False
                    
        except Exception as e:
            logger.error(f"Database update failed: {e}")
            logger.error(f"Error type: {type(e).__name__}")
            # Store the error for detailed reporting
            self.last_db_error = f"{type(e).__name__}: {str(e)}"
            return False
    
    def process_qtest_upload(self, file_content: bytes, filename: str, comparison_id: int) -> Dict:
        """
        Complete workflow: validate, upload to blob, update database
        
        Returns:
            Result dictionary with status and details
        """
        result = {
            "success": False,
            "comparison_id": comparison_id,
            "validation": None,
            "blob_url": None,
            "database_updated": False,
            "error": None
        }
        
        try:
            # Step 1: Validate file
            is_valid, validation_details = self.validate_qtest_file(file_content, filename)
            result["validation"] = validation_details
            
            if not is_valid:
                result["error"] = "Validation failed"
                return result
            
            # Step 2: Upload to Azure Blob
            blob_url = self.upload_to_azure_blob(file_content, filename, comparison_id)
            
            if not blob_url:
                # Include specific error details
                specific_error = getattr(self, 'last_upload_error', 'Unknown Azure upload error')
                result["error"] = f"Failed to upload to Azure Blob Storage: {specific_error}"
                return result
            
            result["blob_url"] = blob_url
            
            # Step 3: Update database
            db_updated = self.update_comparison_qtest_url(comparison_id, blob_url)
            
            if not db_updated:
                # Try to delete blob if database update fails
                try:
                    blob_client = self.blob_service_client.get_blob_client(
                        container=self.container_name,
                        blob=blob_url.split(f"/{self.container_name}/")[-1]
                    )
                    blob_client.delete_blob()
                    logger.info("Rolled back blob upload due to database failure")
                except:
                    logger.error("Failed to rollback blob upload")
                
                # Include specific database error if available
                db_error = getattr(self, 'last_db_error', 'Unknown database error')
                result["error"] = f"Failed to update database: {db_error}"
                return result
            
            result["database_updated"] = True
            result["success"] = True
            
            logger.info(f"Successfully processed QTest upload for comparison {comparison_id}")
            
        except Exception as e:
            logger.error(f"Error processing QTest upload: {e}")
            result["error"] = str(e)
        
        return result